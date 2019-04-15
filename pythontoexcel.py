#Python to Excel

import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

MySheet = wb.active

MySheet.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')

MySheet['A1'] = 'An example of Sum Formula'

MySheet['B2'] = 50
MySheet['B3'] = 50
MySheet['B4'] = 50

MySheet['A5'] = 'Total'
MySheet['B8'] = '=SUM(B2:B7)'

MySheet.column_dimensions['A'].width = 25

write_sheet = wb['Second Sheet']

read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row

write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

write_row = 2
write_col_A = 1
write_col_B = 2
write_col_C = 3
write_col_D = 4


for row in read_ws.iter_rows(min_row = 2, max_row = maxR, max_col = maxC):
    #Read the data
    name = row[0].value
    cost = float(row[1].value)
    amt_sold = float(row[2].value)
    total = float(row[3].value)

    #Write the data
    write_sheet.cell(write_row, write_col_A).value = name
    write_sheet.cell(write_row, write_col_B).value = cost
    write_sheet.cell(write_row, write_col_C).value = amt_sold
    write_sheet.cell(write_row, write_col_D).value = total

    write_row += 1


total_row = str(write_row+2)
write_sheet['A'+total_row] = 'Total:'
write_sheet['B'+total_row] = '=SUM(B2:B' + str(write_row) + ')'
write_sheet['C'+total_row] = '=SUM(C2:C' + str(write_row) + ')'
write_sheet['D'+total_row] = '=SUM(D2:D' + str(write_row) + ')'



wb.save('PythonToExcel.xlsx')
